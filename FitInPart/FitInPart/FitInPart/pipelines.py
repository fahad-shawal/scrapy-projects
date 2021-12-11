from openpyxl import Workbook
from itemadapter import ItemAdapter



class FitinpartPipeline:

    def open_spider(self, spider):
        spider.wb = Workbook()
        wb = spider.wb
        parts = wb.create_sheet("Parts")
        parts.append(['part_id', 'url', 'name', 'brand', 'image_urls', 'description'])
        
        prices = wb.create_sheet("Prices")
        prices.append(['part_id', 'vendor', 'price', 'days', 'qty'])

        vehicles = wb.create_sheet("Vehicles")
        vehicles.append(['part_id', 'brand', 'model', 'placement', 'production', 'application', 'body', 'eng', 'note'])

        oe_number = wb.create_sheet("Oe_Number")
        oe_number.append(['part_id', 'owner', 'part_number'])

        sub_assembly = wb.create_sheet("Sub_assembly")
        sub_assembly.append(['part_id', 'brand', 'part', 'type'])

        specifications = wb.create_sheet("Specifications")
        specifications.append(['part_id', 'Location', 'Height', 'Length-1', 'Pcs In Set', 'Thickness-1', 
        'Width-1', 'O-Ring gasket','Structure Material', 'Thread', 'type', 'Valves', 'Out'] )
        
        parent_assembly = wb.create_sheet("Parent_assembly")
        parent_assembly.append(['part_id', 'brand', 'part', 'type'])

        filters = wb.create_sheet("Filters")
        filters.append(['part_id', 'brand', 'model', 'years', 'body', 'engine', 'engin_no'])

        spider.wb.save("data.xlsx")

    def close_spider(self, spider):
        spider.wb.save('data.xlsx')    

    def process_item(self, item, spider):
        spider.wb['Parts'].append([
            item['part_id'], item['url'], item['name'], item['brand'], item['image_urls'], item['description']
        ])

        for price in item['prices']:
            spider.wb['Prices'].append([
                item['part_id'], price['vendor'], price['price'], price['days'], price['qty']
            ])

        for eo_no in item['oe_number']:
            spider.wb['Oe_Number'].append([
                item['part_id'], eo_no['owner'], eo_no['part_number']
            ])

        for vehicle in item['vehicles']:
            spider.wb['Vehicles'].append([
                item['part_id'], vehicle['brand'], vehicle['model'], vehicle['placement'], 
                vehicle['production'], vehicle['application'], vehicle['body'], vehicle['eng'], 
                vehicle['note']
            ])

        for sub_assembly in item['sub_assembly']:
            spider.wb['Sub_assembly'].append([
                item['part_id'], sub_assembly['brand'], sub_assembly['part'], sub_assembly['type']
            ])

        for par_assembly in item['parent_assembly']:
            spider.wb['Parent_assembly'].append([
                item['part_id'], par_assembly['brand'], par_assembly['part'], par_assembly['type']
            ])

        spider.wb['Specifications'].append([
            item['part_id'], item['specifications']['location'], item['specifications']['height'], 
            item['specifications']['length-1'], item['specifications']['pcs in set'], 
            item['specifications']['thickness-1'], item['specifications']['width-1'], 
            item['specifications']['o-ring gasket'],item['specifications']['structure material'], 
            item['specifications']['thread'], item['specifications']['type'], item['specifications']['valves'], 
            item['specifications']['out']
        ])
        
        spider.wb['Filters'].append([
            item['part_id'], item['filters']['brand'], item['filters']['model'], item['filters']['years'], 
            item['filters']['body'], item['filters']['engine'], item['filters']['engin_no']
        ])

        spider.wb.save("data.xlsx")
        return item
